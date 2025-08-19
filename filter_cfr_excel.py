# -*- coding: utf-8 -*-
import pandas as pd
import argparse
import os
import requests
import json
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook

URL_EXCEL = "https://servicio.pesca.mapama.es/CENSO/ConsultaBuqueRegistro/Buques/DownloadPesquerosReport"
LOCAL_INPUT = "data/Listado_buques_pesca.xlsx"

def parse_args():
    parser = argparse.ArgumentParser(description="Descarga y filtra un Excel por CFR y Estado.")
    parser.add_argument("--output", required=True, help="Archivo Excel de salida")
    parser.add_argument("--jsonld", help="Archivo JSON-LD de salida")
    parser.add_argument("--rdf", help="Archivo RDF de salida")
    parser.add_argument("--desde", required=True, help="CFR desde")
    parser.add_argument("--hasta", required=True, help="CFR hasta")
    parser.add_argument("--estado", help="Filtrar por estado (opcional, coma separada)")
    return parser.parse_args()

def descargar_excel(url, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    response = requests.get(url)
    response.raise_for_status()
    with open(output_path, "wb") as f:
        f.write(response.content)

def exportar_estados(df, path="output/estados.json"):
    estados = df["Estado"].dropna().unique().tolist()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(estados, f, ensure_ascii=False, indent=2)

def aplicar_autofiltro_y_estilo(ws, num_cols):
    col_final = chr(64 + num_cols) if num_cols <= 26 else 'Z'  # Simplificación básica
    ws.auto_filter.ref = "A2:{}2".format(col_final)

    # Estilo para encabezados
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill

def main():
    args = parse_args()
    descargar_excel(URL_EXCEL, LOCAL_INPUT)

    df = pd.read_excel(LOCAL_INPUT, header=3)
    exportar_estados(df)  # Genera archivo para frontend

    df_filtrado = df[(df["CFR"] >= args.desde) & (df["CFR"] <= args.hasta)]

    if args.estado:
        estados_filtrados = [e.strip() for e in args.estado.split(",")]
        df_filtrado = df_filtrado[df_filtrado["Estado"].isin(estados_filtrados)]

    columnas = [
        "CFR", "Nombre", "Matrícula", "Estado", "Eslora total (m)",
        "Arqueo GT", "Potencia kW", "Material del casco",
        "Puerto base", "Censo por modalidad"
    ]
    columnas_validas = [col for col in columnas if col in df_filtrado.columns]
    df_final = df_filtrado[columnas_validas]

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    with pd.ExcelWriter(args.output, engine='openpyxl') as writer:
        df_final.to_excel(writer, startrow=1, index=False)
        ws = writer.sheets['Sheet1']

        # Mensaje informativo en la fila 1
        mensaje = "Filtro aplicado: CFR desde {} hasta {}".format(args.desde, args.hasta)
        ws.cell(row=1, column=1, value=mensaje)

        # Aplicar autofiltro y estilo a encabezados
        aplicar_autofiltro_y_estilo(ws, len(columnas_validas))

    if args.jsonld:
        df_final.to_json(args.jsonld, orient="records", force_ascii=False, indent=2)

    if args.rdf:
        with open(args.rdf, "w", encoding="utf-8") as f:
            for _, row in df_final.iterrows():
                f.write("<http://example.org/buque/{}> a <http://example.org/Buque> .\n".format(row['CFR']))

if __name__ == "__main__":
    main()
